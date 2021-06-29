import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Border, Side,Alignment

# Error handling for the entire program
try:
    # declaring file paths

    print("\n")
    csv_name = str(input("Enter the name of Zoom participant list (csv file) file :\n"))
    csv_path = csv_name + ".csv"
    print("\n")
    file_cs = pd.read_csv(csv_path, skiprows=1)

    class_name = str(input("Enter the name of the class (e.g. 1st year):\n"))
    excel_name = "Attendance Sheet.xlsx"
    print("\n")
    file_ex = pd.read_excel(excel_name, sheet_name=class_name, skiprows=1)
    file_report = 'Report.xlsx'

    # initializing necessary values
    now = datetime.now()
    dt = now.strftime("%d/%m/%Y (%I:%M %p)")
    count_cs = 0
    count_ex = 0
    count_unmatched = 0
    style = Border(bottom=Side(color='00000000',style="thin"))
    style_cell= Alignment(horizontal="center")

    # introduction of column number
    print("Attendance column is titled by the date on which classes are held.")
    print("When the program is run for the first time, the column number of the attendance column will be 4. Next, it "
          "will "
          "be 5 and gradually it "
          "will "
          "increase by 1.\n")
    print("*** The previous column number can be found in the last column of the "
          "Report File *** "
          "\n")
    column_number = int(input("Enter a new column number( next number of the previous column number ) in which "
                              "attendance will be "
                              "recorded:\n"))
    while column_number < 4:
        print("!! Sorry, invalid column number, column number should be 4 or greater than 4 !!")
        column_number = int(input("Enter a new column number( next number of the previous column number ) in which "
                                  "attendance will be "
                                  "recorded:\n"))
    print("\n")
    number_Present_Students = 0
    unmatched_name = dict()
    unmatched_mail = dict()

    # declare data frame
    df_cs = pd.DataFrame(file_cs)
    df_ex = pd.DataFrame(file_ex)
    Row_ex = len(df_ex)
    Row_cs = len(df_cs)
    # define workbook
    wb = openpyxl.load_workbook(excel_name)
    ws = wb[class_name]

    # loop for application
    while count_cs in range(Row_cs):
        if count_ex in range(Row_ex):
            if (df_cs.iloc[count_cs][0] == df_ex.iloc[count_ex][1]) and (
                    df_cs.iloc[count_cs][1] == df_ex.iloc[count_ex][2]):
                # give attendance to matched values
                ws.cell(row=count_ex + 3, column=column_number, value=1)
                number_Present_Students += 1
                wb.save(excel_name)
                wb.close()
                count_cs += 1
                count_ex = 0
            else:
                # if values don't match
                count_ex += 1
        else:
            # unmatched values after scanning the full excel file
            unmatched_name[count_unmatched] = df_cs.iloc[count_cs][0]
            unmatched_mail[count_unmatched] = df_cs.iloc[count_cs][1]
            count_unmatched += 1
            count_ex = 0
            count_cs += 1
    # write report of unmatched values
    wb_report = openpyxl.load_workbook(file_report)
    ws_report = wb_report[class_name]
    for col_report in range(1, 7, 1):
        last_row = ws_report.max_row
        if col_report == 1:
            c=ws_report.cell(row=ws_report.max_row + 1, column=col_report, value=class_name)
            c.alignment=style_cell
        elif col_report == 2:
            c0=ws_report.cell(row=ws_report.max_row, column=col_report, value=dt)
            c0.alignment=style_cell
        elif col_report == 3:
            count_unmatched-=1
            while count_unmatched >= 0:
                name = unmatched_name.get(count_unmatched)
                mail = unmatched_mail.get(count_unmatched)
                cc=ws_report.cell(row=last_row, column=col_report, value=name)
                cc1=ws_report.cell(row=last_row, column=4, value=mail)
                cc.alignment=style_cell
                cc1.alignment=style_cell
                last_row += 1
                count_unmatched -= 1
            # bordering cells
            c3 = ws_report.cell(row=last_row-1, column=3)
            c4 = ws_report.cell(row=last_row-1, column=4)
            c3.border = style
            c4.border = style
        elif col_report == 5:
            c5 = ws_report.cell(row=ws_report.max_row, column=col_report, value=number_Present_Students)
            c5.border = style
            c5.alignment=style_cell
        elif col_report == 6:
            c6 = ws_report.cell(row=ws_report.max_row, column=col_report, value=column_number)
            c6.border = style
            c6.alignment=style_cell
            c1=ws_report.cell(row=ws_report.max_row, column=1)
            c2=ws_report.cell(row=ws_report.max_row,column=2)
            c1.border=style
            c2.border=style
    wb_report.save(file_report)
    wb_report.close()

    print("Please open the Attendance Sheet file")
    print("For more details open the Report file\n")
    print("Your work is finished. Exit Program\n")

except(RuntimeError, OverflowError, RuntimeWarning, InterruptedError, TypeError, NameError,ValueError):
    print("\n")
    exit("An unexpected error occurred! Exit the program and run again")
except PermissionError:
    print("Close all files and restart the program")
    exit()
except FileNotFoundError:
    print("\n")
    exit("Error: File not found! Check if the directory is same")
